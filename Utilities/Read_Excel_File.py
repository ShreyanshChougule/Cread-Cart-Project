import openpyxl


class Read_Excel_File:

    @staticmethod
    def read_excel(rownum, colnum):
        Excel = openpyxl.load_workbook("C:\\Users\\Tejas\\Class Project\\TestData\\Test_Data.xlsx")
        Sheet = Excel.active
        return Sheet.cell(row=rownum, column=colnum).value

    @staticmethod
    def write_excel(rownum, colnum, data):
        Book = openpyxl.load_workbook("C:\\Users\\Tejas\\Class Project\\TestData\\Test_Data.xlsx")
        Sheet = Book.active
        Sheet.cell(row=rownum, column=colnum).value = data
        Book.save("C:\\Users\\Tejas\\Class Project\\TestData\\Test_Data.xlsx")